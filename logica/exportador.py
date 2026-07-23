from openpyxl import load_workbook
from pathlib import Path

class Exportador:

    def __init__(self):

        self.ruta = None
        self.libro = None
        self.hoja = None
        self.fila_actual = None
        self.abierto = False

    #----escoge si quieres crear un excel    
    def crear_excel(self, ruta):
        self.ruta = ruta

    #----este es para une excel que ya existe 
    def abrir_excel(self, ruta):

        self.ruta = ruta

        self.libro = load_workbook(ruta)

        self.hoja = self.libro.active

        self.fila_actual = self.hoja.max_row + 1

        return True


    #----en caso de que cojas uno que ya existe, te busca donde escribir 
    def buscar_siguiente_fila(
        self.fila_actual = self.hoja.max_row + 1
    )
    
    
    #----guarda el resultado con las columnas que queremos 
    def guardar_resultado(
        self,
        trazabilidad,
        resultado,
        referencia=None,
        caja=None, 
        etiqueta=None
    )

    def guardar(
        self.libro.save(self.ruta)
    )
        



    